# -*- coding: UTF-8 -*-

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import sv_ttk
import database
from ctypes import windll
import help
from tkcalendar import DateEntry
import locale
from datetime import datetime
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
import shutil
import os


class Orders:
    def __init__(self, parent, products_list, customers_list, b_frame):
        self.b_frame = b_frame
        self.cart = []
        self.customer_name_e = None
        self.date_e = None
        self.product_name_e = None
        self.product_qtt_e = None
        self.is_payed = None
        self.products = products_list
        self.products_tmp = [list(t) for t in products_list]
        self.customers = customers_list
        self.my_tree = None
        self.parent = parent
        self.frame = ttk.Frame(parent)
        self.setup_ui()

    def setup_ui(self):
        data_frame = ttk.LabelFrame(self.frame, text="", labelanchor=tk.NE)
        data_frame.pack(fill="x", expand=True, padx=20, pady=5)

        data_frame.grid_columnconfigure(0, weight=1)
        data_frame.grid_columnconfigure(1, weight=1)
        data_frame.grid_columnconfigure(2, weight=1)
        data_frame.grid_columnconfigure(3, weight=1)

        # Data entry
        customer_name_l = ttk.Label(data_frame, text="اسم الزبون")
        self.customer_name_e = ttk.Combobox(data_frame, font=normal_text, justify="right",
                                            values=[f"({c[5]}) {c[1]} {c[2]}" for c in self.customers],
                                            state="readonly")
        customer_name_l.grid(row=0, column=4, padx=10, pady=10)
        self.customer_name_e.grid(row=0, column=3, padx=10, pady=10, sticky="ew")

        date_l = ttk.Label(data_frame, text="تاريخ الطلب")
        self.date_e = DateEntry(data_frame, justify="center", state="readonly", background='black', foreground='white',
                                borderwidth=1,
                                font=normal_text, showweeknumbers=False, showothermonthdays=False,
                                selectbackground='red',
                                date_pattern="YYYY-mm-dd", firstweekday="sunday", locale="ar_DZ", weekenddays=[6, 7],
                                )
        date_l.grid(row=0, column=2, padx=10, pady=10)
        self.date_e.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        self.is_payed = ttk.Combobox(data_frame, font=normal_text, justify="right", state="readonly",
                                     values=["لم يتم دفع الفاتورة", "تم دفع الفاتورة"])
        self.is_payed.set("لم يتم دفع الفاتورة")
        self.is_payed.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        add2cart_btn = ttk.Button(data_frame, text="أضف إلى السلة", style="success.TButton",
                                  command=self.add2cart)
        add2cart_btn.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

        product_qtt_l = ttk.Label(data_frame, text="الكمية")
        self.product_qtt_e = ttk.Entry(data_frame, font=normal_text, justify="right")
        product_qtt_l.grid(row=2, column=2, padx=10, pady=10)
        self.product_qtt_e.grid(row=2, column=1, padx=10, pady=10, sticky="ew")

        product_name_l = ttk.Label(data_frame, text="السلعة")
        self.product_name_e = ttk.Combobox(data_frame, font=normal_text, justify="right",
                                           values=[f"(#{p[0]}) {p[1]} {p[2]}" for p in self.products], state="readonly")
        product_name_l.grid(row=2, column=4, padx=10, pady=10)
        self.product_name_e.grid(row=2, column=3, padx=10, pady=10, sticky="ew")

        add2order = ttk.Button(data_frame, text="تأكيد الطلب", style="danger.TButton", command=self.confirm_order)
        add2order.grid(columnspan=5, padx=500, pady=30, sticky="ew")

        tree_frame = ttk.Frame(self.frame)
        tree_frame.pack(fill="x", expand=True, padx=20, pady=5)

        self.my_tree = ttk.Treeview(tree_frame, selectmode="extended", height=5)
        self.my_tree.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        scroll_frame = ttk.Frame(tree_frame)
        scroll_frame.pack(side=tk.RIGHT, fill=tk.Y)

        tree_scroll = ttk.Scrollbar(scroll_frame, orient="vertical", command=self.my_tree.yview)
        tree_scroll.pack(side=tk.LEFT, fill="y")

        self.my_tree.config(yscrollcommand=tree_scroll.set)

        cols = ("price_tax", "price_tot", "price", "qtt", "product", "id")
        self.my_tree['columns'] = cols

        self.my_tree.column("#0", width=0, stretch=tk.NO)
        self.my_tree.column("id", width=0, stretch=tk.NO)
        self.my_tree.column("product", anchor=tk.CENTER)
        self.my_tree.column("qtt", anchor=tk.CENTER)
        self.my_tree.column("price", anchor=tk.CENTER)
        self.my_tree.column("price_tot", anchor=tk.CENTER)
        self.my_tree.column("price_tax", anchor=tk.CENTER)

        self.my_tree.heading("#0", text="", anchor=tk.W)
        self.my_tree.heading("product", anchor=tk.CENTER, text="السلعة")
        self.my_tree.heading("qtt", anchor=tk.CENTER, text="الكمية")
        self.my_tree.heading("price", anchor=tk.CENTER, text="سعر الوحدة")
        self.my_tree.heading("price_tot", anchor=tk.CENTER, text="المبلغ الصافي")
        self.my_tree.heading("price_tax", anchor=tk.CENTER, text="المبلغ الاجمالي")
        self.my_tree.bind('<Delete>',
                          lambda event: self.delete_row()
                          if self.my_tree.identify_region(event.x, event.y) != "heading" else None)

    def delete_row(self):
        selected_items = self.my_tree.selection()
        for selected_item in selected_items:
            values = self.my_tree.item(selected_item, "values")
            self.cart = [item for item in self.cart if int(item["id"]) != int(values[5])]
            for item in self.products_tmp:
                if int(item[0]) == int(values[5]):
                    item[6] = item[6] + float(values[3])
            self.my_tree.delete(selected_item)

    def validate(self):
        errors = []
        if not self.product_name_e.get():
            errors.append("يجب عليك اختيار سلعة")
        if not help.is_float(self.product_qtt_e.get()) or float(self.product_qtt_e.get()) == 0:
            errors.append("تحقق من الكمية المطلوبة")
        return errors

    def add2cart(self):
        errors = self.validate()
        if errors:
            messagebox.showerror("خطأ", "\n".join(errors), parent=self.parent)
        else:
            product_id = int(
                self.product_name_e.get()[self.product_name_e.get().find("#") + 1:self.product_name_e.get().find(")")])
            selected_product = [p for p in self.products_tmp if p[0] == product_id][0]
            if selected_product[6] < float(self.product_qtt_e.get()):
                messagebox.showwarning("الكمية كبيرة",
                "الكمية التي أدخلتها غير متوفرة .. يرجى تحديث الكمية المتوفرة الخاصة بالسلعة قبل القيام بأي عملية")
            else:
                for item in self.products_tmp:
                    if item[0] == product_id:
                        item[6] = item[6] - float(self.product_qtt_e.get())
                # Check if any item in the cart has the same ID as the selected product
                if any(item["id"] == selected_product[0] for item in self.cart):
                    # If yes, update the quantity by adding the new quantity
                    for item in self.cart:
                        if item["id"] == selected_product[0]:
                            item["quantity"] += float(self.product_qtt_e.get())
                            break
                else:
                    # If no, add a new item to the cart
                    self.cart.append({
                        "id": int(selected_product[0]),
                        "product_name": selected_product[1],
                        "product_category": selected_product[2],
                        "quantity": float(self.product_qtt_e.get()),
                        "price_at_order_time": float(selected_product[3]),
                        "vat_tax_at_order_time": float(selected_product[4]),
                        "stamp_tax_at_order_time": True if selected_product[5] == "نعم" else False
                    })

                self.my_tree.delete(*self.my_tree.get_children())

                for item in self.cart:
                    price = item["quantity"] * item["price_at_order_time"]
                    price_tax = price + (price * item["vat_tax_at_order_time"] / 100)
                    price_tax = price_tax + (price / 100) if item["stamp_tax_at_order_time"] is True else price_tax
                    self.my_tree.insert(parent='', index='end', text='',
                                        values=(price_tax, price, item["price_at_order_time"], item["quantity"],
                                                f"{item["product_name"]} - {item["product_category"]}", item["id"]))

    def clean_data(self):
        self.my_tree.delete(*self.my_tree.get_children())
        self.cart = []
        self.customer_name_e.set('')
        self.product_name_e.set('')
        self.date_e.set_date(datetime.today().strftime('%Y-%m-%d'))
        self.is_payed.current(0)
        self.product_qtt_e.delete(0, tk.END)

    def update_products(self, new_values):
        self.products = new_values
        self.products_tmp = [list(t) for t in new_values]
        self.clean_data()
        self.product_name_e["values"] = [f"(#{p[0]}) {p[1]} - {p[2]}" for p in self.products]

    def update_customers(self, new_values):
        self.customers = new_values
        self.customer_name_e["values"] = [f"({c[5]}) {c[1]} {c[2]}" for c in self.customers]

    def confirm_order(self):
        if len(self.cart) < 1:
            messagebox.showerror("خطأ", "أضف سلعا قبل تأكيد الطلب", parent=self.parent)
        elif len(self.customer_name_e.get()) == 0:
            messagebox.showerror("خطأ", "اختر زبونا قبل تأكيد الطلب", parent=self.parent)
        else:
            msg = "هل أنت متأكد أن معلومات الزبون و تاريخ الطلب و معلومات السلع صحيحة ؟"
            if messagebox.askyesno("تأكيد", msg, parent=self.parent) is True:
                customer_reg_id = self.customer_name_e.get()[
                                  self.customer_name_e.get().find("(") + 1:self.customer_name_e.get().find(")")]
                selected_customer = [c for c in self.customers if c[5] == customer_reg_id][0]
                order = {
                    "customer_id": selected_customer[0],
                    "order_date": self.date_e.get(),
                    "order_status": self.is_payed.get(),
                    "cart": self.cart
                }
                print(order)
                # insert to database and modify the product database
                msg = db.add_order(order)
                if msg.startswith("تم"):
                    messagebox.showinfo("", msg)
                else:
                    messagebox.showerror("", msg)
                # update the gui for the product frame, and bills frame
                msg2 = db.update_product_quantity(self.cart)
                self.update_products(db.get_products())
                # Display success message and clean the data
                self.clean_data()


class Products:
    def __init__(self, parent, data, o_frame):
        self.o_frame = o_frame
        self.product_name_e = None
        self.product_category_e = None
        self.product_price_e = None
        self.product_unit_e = None
        self.product_quantity_e = None
        self.product_tva_e = None
        self.product_stamp_e = None
        self.data = data
        self.my_tree = None
        self.selected_product = None
        self.parent = parent
        self.frame = ttk.Frame(parent)
        self.setup_ui()

    def setup_ui(self):
        tree_frame = ttk.Frame(self.frame)
        tree_frame.pack(fill=tk.X, expand=True, padx=20, pady=5)

        self.my_tree = ttk.Treeview(tree_frame, selectmode="extended", height=5)
        self.my_tree.pack(side=tk.RIGHT, fill=tk.X, expand=True)

        scroll_frame = ttk.Frame(tree_frame)
        scroll_frame.pack(side=tk.RIGHT, fill=tk.Y)

        tree_scroll = ttk.Scrollbar(scroll_frame, orient="vertical", command=self.my_tree.yview)
        tree_scroll.pack(side=tk.LEFT, fill="y")

        self.my_tree.config(yscrollcommand=tree_scroll.set)

        cols = ("stamp_tax", "vat_tax", "price", "quantity", "category", "product", "id")
        self.my_tree['columns'] = cols

        self.my_tree.column("#0", width=0, stretch=tk.NO)
        self.my_tree.column("id", width=0, stretch=tk.NO)
        self.my_tree.column("product", anchor=tk.CENTER)
        self.my_tree.column("category", anchor=tk.CENTER)
        self.my_tree.column("quantity", anchor=tk.CENTER)
        self.my_tree.column("price", anchor=tk.CENTER)
        self.my_tree.column("vat_tax", anchor=tk.CENTER)
        self.my_tree.column("stamp_tax", anchor=tk.CENTER)

        self.my_tree.heading("#0", text="", anchor=tk.W)
        self.my_tree.heading("product", anchor=tk.CENTER, text="السلعة")
        self.my_tree.heading("category", anchor=tk.CENTER, text="النوع")
        self.my_tree.heading("price", anchor=tk.CENTER, text="السعر")
        self.my_tree.heading("quantity", anchor=tk.CENTER, text="الكمية")
        self.my_tree.heading("vat_tax", anchor=tk.CENTER, text="الضريبة المضافة")
        self.my_tree.heading("stamp_tax", anchor=tk.CENTER, text="ضريبة الطابع")
        self.my_tree.bind('<Double-1>', lambda event: self.select_product() if self.my_tree.identify_region(
            event.x, event.y) != "heading" else None)

        for record in self.data:
            self.my_tree.insert(parent='', index='end', text='',
                                values=(record[5], f"% {record[4]}", f"{record[3]}",
                                        f"{record[6]:.2f} {record[7]}", record[2], record[1], record[0]))

        data_frame = ttk.LabelFrame(self.frame, text="", labelanchor=tk.NE)
        data_frame.pack(fill="x", expand=True, padx=20, pady=5)

        data_frame.grid_columnconfigure(0, weight=1)
        data_frame.grid_columnconfigure(1, weight=1)
        data_frame.grid_columnconfigure(2, weight=1)
        data_frame.grid_columnconfigure(3, weight=1)

        product_name_l = ttk.Label(data_frame, text="السلعة")
        self.product_name_e = ttk.Entry(data_frame, font=normal_text, justify="right")
        product_name_l.grid(row=0, column=3, padx=10, pady=10)
        self.product_name_e.grid(row=0, column=2, padx=10, pady=10, sticky="ew")

        product_category_l = ttk.Label(data_frame, text="النوع")
        self.product_category_e = ttk.Entry(data_frame, font=normal_text, justify="right")
        product_category_l.grid(row=1, column=3, padx=10, pady=10)
        self.product_category_e.grid(row=1, column=2, padx=10, pady=10, sticky="ew")

        product_price_l = ttk.Label(data_frame, text="السعر (دج)")
        self.product_price_e = ttk.Entry(data_frame, font=normal_text, justify="right")
        product_price_l.grid(row=1, column=1, padx=10, pady=10)
        self.product_price_e.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

        product_quantity_l = ttk.Label(data_frame, text="الكمية")
        self.product_quantity_e = ttk.Entry(data_frame, font=normal_text, justify="right")
        product_quantity_l.grid(row=2, column=3, padx=10, pady=10)
        self.product_quantity_e.grid(row=2, column=2, padx=10, pady=10, sticky="ew")

        product_unit_l = ttk.Label(data_frame, text="وحدة القياس")
        self.product_unit_e = ttk.Entry(data_frame, font=normal_text, justify="right")
        self.product_unit_e.insert(0, "كغ")
        product_unit_l.grid(row=2, column=1, padx=10, pady=10)
        self.product_unit_e.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

        product_tva_l = ttk.Label(data_frame, text="(%) الضريبة المضافة")
        self.product_tva_e = ttk.Combobox(data_frame, font=normal_text, justify="right", values=["0", "9", "19"])
        self.product_tva_e.current(0)
        product_tva_l.grid(row=3, column=3, padx=10, pady=10)
        self.product_tva_e.grid(row=3, column=2, padx=10, pady=10, sticky="ew")

        product_stamp_l = ttk.Label(data_frame, text="ضريبة الطابع")
        self.product_stamp_e = ttk.Combobox(data_frame, font=normal_text, justify="right", values=["نعم", "لا"],
                                            state='readonly')
        self.product_stamp_e.current(1)
        product_stamp_l.grid(row=3, column=1, padx=10, pady=10)
        self.product_stamp_e.grid(row=3, column=0, padx=10, pady=10, sticky="ew")

        buttons_frame = ttk.LabelFrame(self.frame, text="", labelanchor=tk.NE)
        buttons_frame.pack(fill="x", expand=tk.YES, padx=20)

        update_button = ttk.Button(buttons_frame, text="تحديث سلعة", style="warning.TButton",
                                   command=self.update_product)
        update_button.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        add_button = ttk.Button(buttons_frame, text="إضافة سلعة", style="success.TButton", command=self.add_product)
        add_button.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        buttons_frame.grid_columnconfigure(0, weight=1)
        buttons_frame.grid_columnconfigure(1, weight=1)

    def update_products_list(self):
        self.my_tree.delete(*self.my_tree.get_children())
        self.data = db.get_products()

        # Insert the updated data into the Treeview
        for product in self.data:
            self.my_tree.insert(parent='', index='end', text='', values=(
                product[5], f"% {product[4]}", product[3], f"{product[6]:.2f} {product[7]}",
                product[2], product[1], product[0]))

    def validation(self):
        errors = []
        if not help.is_arabic(self.product_name_e.get().strip()):
            errors.append("اسم السلعة يجب أن يكون باللغة العربية")
        if not help.is_float(self.product_price_e.get().strip()):
            errors.append("تحقق من أن السعر صحيح")
        if not help.is_float(self.product_quantity_e.get().strip()) or float(self.product_quantity_e.get()) == 0:
            errors.append("تحقق من أن الكمية صحيحة")
        if not help.is_arabic(self.product_unit_e.get().strip()):
            errors.append("وحدة القياس يجب أن تكون باللغة العربية")
        if not help.is_float(self.product_tva_e.get().strip()):
            errors.append("الضريبة المضافة يجب أن تكون عدد حقيقي")
        return errors

    def clear_entries(self):
        self.selected_product = None
        self.product_name_e.delete(0, tk.END)
        self.product_category_e.delete(0, tk.END)
        self.product_quantity_e.delete(0, tk.END)
        self.product_price_e.delete(0, tk.END)
        self.product_unit_e.delete(0, tk.END)
        self.product_unit_e.insert(0, "كغ")
        self.product_tva_e.set("0")
        self.product_stamp_e.set("لا")

    def select_product(self):
        values = self.my_tree.item(self.my_tree.focus(), 'values')
        self.clear_entries()
        self.selected_product = values[6]
        self.product_name_e.insert(0, values[5])
        self.product_category_e.insert(0, values[4])
        self.product_quantity_e.insert(0, values[3].split(' ')[0])
        self.product_price_e.insert(0, values[2])
        self.product_tva_e.set(values[1].split(' ')[1])
        self.product_stamp_e.set(values[0])

    def add_product(self):
        errors = self.validation()
        if errors:
            messagebox.showerror("خطأ", "\n".join(errors), parent=self.parent)
        else:
            if any(product[1] == self.product_name_e.get().strip() and product[2] == self.product_category_e.get().
                    strip() for product in self.data):
                messagebox.showwarning("تحذير", "السلعة بالاسم والنوع المدخلين موجودة بالفعل في قاعدة البيانات",
                                       parent=self.parent)
            else:
                data = {
                    "name": self.product_name_e.get().strip(),
                    "category": self.product_category_e.get().strip(),
                    "price": float(self.product_price_e.get().strip()),
                    "vat_tax": float(self.product_tva_e.get().strip()),
                    "stamp_tax": self.product_stamp_e.get().strip(),
                    "available_quantity": float(self.product_quantity_e.get().strip()),
                    "unit": self.product_unit_e.get().strip()
                }
                msg = db.add_product(data)
                if msg.startswith("تم"):
                    self.update_products_list()
                    self.clear_entries()
                    messagebox.showinfo("", msg)
                else:
                    messagebox.showerror("", msg)

    def update_product(self):
        errors = self.validation()
        if errors:
            messagebox.showerror("خطأ", "\n".join(errors), parent=self.parent)
        else:
            if any(product[1] == self.product_name_e.get().strip() and product[2] == self.product_category_e.get().
                    strip() and str(product[0]) != self.selected_product for product in self.data):
                messagebox.showwarning("تحذير", "السلعة بالاسم والنوع المدخلين موجودة بالفعل في قاعدة البيانات",
                                       parent=self.parent)
            elif any(str(product[0]) == self.selected_product for product in self.data):
                data = {
                    "name": self.product_name_e.get().strip(),
                    "category": self.product_category_e.get().strip(),
                    "price": float(self.product_price_e.get().strip()),
                    "vat_tax": float(self.product_tva_e.get().strip()),
                    "stamp_tax": self.product_stamp_e.get().strip(),
                    "available_quantity": float(self.product_quantity_e.get().strip()),
                    "unit": self.product_unit_e.get().strip(),
                    "id": self.selected_product.strip()
                }
                msg = db.update_product(data)
                if msg.startswith("تم"):
                    self.update_products_list()
                    self.clear_entries()
                    messagebox.showinfo("", msg)
                else:
                    messagebox.showerror("", msg)
            else:
                messagebox.showerror("خطأ", "هذه السلعة لا توجد في قاعدة البيانات", parent=self.parent)


class Customers:
    def __init__(self, parent, data, o_frame):
        self.o_frame = o_frame
        self.customer_name_e = None
        self.customer_lastname_e = None
        self.customer_phone_e = None
        self.customer_address_e = None
        self.customer_reg_id_e = None
        self.customer_tax_item_e = None
        self.customer_tax_num_e = None
        self.customer_statistical_id_e = None
        self.data = data
        self.my_tree = None
        self.selected_customer = None
        self.parent = parent
        self.frame = ttk.Frame(parent)
        self.setup_ui()

    def setup_ui(self):
        tree_frame = ttk.Frame(self.frame)
        tree_frame.pack(fill=tk.X, expand=True, padx=20, pady=5)

        self.my_tree = ttk.Treeview(tree_frame, selectmode="extended", height=5)
        self.my_tree.pack(side=tk.RIGHT, fill=tk.X, expand=True)

        scroll_frame = ttk.Frame(tree_frame)
        scroll_frame.pack(side=tk.RIGHT, fill=tk.Y)

        tree_scroll = ttk.Scrollbar(scroll_frame, orient="vertical", command=self.my_tree.yview)
        tree_scroll.pack(side=tk.LEFT, fill="y")

        self.my_tree.config(yscrollcommand=tree_scroll.set)

        cols = (
            "stat_id", "tax_item_num", "tax_num", "commercial_register_number", "address", "phone", "lname", "name",
            "id")
        self.my_tree['columns'] = cols

        self.my_tree.column("#0", width=0, stretch=tk.NO)
        self.my_tree.column("id", width=0, stretch=tk.NO)
        self.my_tree.column("name", anchor=tk.CENTER)
        self.my_tree.column("lname", anchor=tk.CENTER)
        self.my_tree.column("phone", anchor=tk.CENTER)
        self.my_tree.column("address", anchor=tk.CENTER)
        self.my_tree.column("commercial_register_number", anchor=tk.CENTER)
        self.my_tree.column("tax_num", width=0, stretch=tk.NO)
        self.my_tree.column("tax_item_num", width=0, stretch=tk.NO)
        self.my_tree.column("stat_id", width=0, stretch=tk.NO)

        self.my_tree.heading("#0", text="", anchor=tk.W)
        self.my_tree.heading("name", anchor=tk.CENTER, text="الاسم")
        self.my_tree.heading("lname", anchor=tk.CENTER, text="اللقب")
        self.my_tree.heading("phone", anchor=tk.CENTER, text="رقم الهاتف")
        self.my_tree.heading("address", anchor=tk.CENTER, text="العنوان")
        self.my_tree.heading("commercial_register_number", anchor=tk.CENTER, text="رقم السجل التجاري")
        self.my_tree.bind('<Double-1>',
                          lambda event: self.select_customer()
                          if self.my_tree.identify_region(event.x, event.y) != "heading" else None)

        for record in self.data:
            self.my_tree.insert(parent='', index='end', text='',
                                values=(record[8], record[7], record[6], record[5], record[4],
                                        record[3] if record[3] != '' else "غير متوفر", record[2], record[1], record[0]))

        data_frame = ttk.LabelFrame(self.frame, text="", labelanchor=tk.NE)
        data_frame.pack(fill="x", expand=True, padx=20, pady=5)

        data_frame.grid_columnconfigure(0, weight=1)
        data_frame.grid_columnconfigure(1, weight=1)
        data_frame.grid_columnconfigure(2, weight=1)
        data_frame.grid_columnconfigure(3, weight=1)

        customer_name_l = ttk.Label(data_frame, text="الاسم")
        self.customer_name_e = ttk.Entry(data_frame, font=normal_text, justify="right")
        customer_name_l.grid(row=0, column=3, padx=10, pady=10)
        self.customer_name_e.grid(row=0, column=2, padx=10, pady=10, sticky="ew")

        customer_lastname_l = ttk.Label(data_frame, text="اللقب")
        self.customer_lastname_e = ttk.Entry(data_frame, font=normal_text, justify="right")
        customer_lastname_l.grid(row=0, column=1, padx=10, pady=10)
        self.customer_lastname_e.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        customer_phone_l = ttk.Label(data_frame, text="رقم الهاتف (اختياري)")
        self.customer_phone_e = ttk.Entry(data_frame, font=normal_text, justify="center")
        customer_phone_l.grid(row=1, column=3, padx=10, pady=10)
        self.customer_phone_e.grid(row=1, column=2, padx=10, pady=10, sticky="ew")

        customer_address_l = ttk.Label(data_frame, text="العنوان")
        self.customer_address_e = ttk.Entry(data_frame, font=normal_text, justify="right")
        customer_address_l.grid(row=1, column=1, padx=10, pady=10)
        self.customer_address_e.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

        customer_reg_id_l = ttk.Label(data_frame, text="رقم السجل (بالفرنسية)")
        self.customer_reg_id_e = ttk.Entry(data_frame, font=normal_text, justify="center")
        customer_reg_id_l.grid(row=2, column=3, padx=10, pady=10)
        self.customer_reg_id_e.grid(row=2, column=2, padx=10, pady=10, sticky="ew")

        customer_tax_num_l = ttk.Label(data_frame, text="الرقم الجبائي")
        self.customer_tax_num_e = ttk.Entry(data_frame, font=normal_text, justify="center")
        customer_tax_num_l.grid(row=2, column=1, padx=10, pady=10)
        self.customer_tax_num_e.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

        customer_tax_item_l = ttk.Label(data_frame, text="رقم مادة الضرائب")
        self.customer_tax_item_e = ttk.Entry(data_frame, font=normal_text, justify="center")
        customer_tax_item_l.grid(row=3, column=3, padx=10, pady=10)
        self.customer_tax_item_e.grid(row=3, column=2, padx=10, pady=10, sticky="ew")

        customer_statistical_id_l = ttk.Label(data_frame, text="رقم التعريف الاحصائي")
        self.customer_statistical_id_e = ttk.Entry(data_frame, font=normal_text, justify="center", state="!invalid")
        customer_statistical_id_l.grid(row=3, column=1, padx=10, pady=10)
        self.customer_statistical_id_e.grid(row=3, column=0, padx=10, pady=10, sticky="ew")

        buttons_frame = ttk.LabelFrame(self.frame, text="", labelanchor=tk.NE)
        buttons_frame.pack(fill="x", expand=tk.YES, padx=20)

        update_button = ttk.Button(buttons_frame, text="تحديث زبون", style="warning.TButton",
                                   command=self.update_customer)
        update_button.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        add_button = ttk.Button(buttons_frame, text="إضافة زبون", style="success.TButton", command=self.add_customer)
        add_button.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        buttons_frame.grid_columnconfigure(0, weight=1)
        buttons_frame.grid_columnconfigure(1, weight=1)

    def validation(self):
        errors = []
        if not help.is_arabic(self.customer_name_e.get().strip()):
            errors.append("اسم الزبون يجب أن يكون باللغة العربية")
        if not help.is_arabic(self.customer_lastname_e.get().strip()):
            errors.append("لقب الزبون يجب أن يكون باللغة العربية")
        if not help.is_phone_number(self.customer_phone_e.get().strip()):
            errors.append("تأكد من أن رقم الهاتف صحيح")
        if not help.is_reg_id(self.customer_reg_id_e.get().strip()):
            errors.append("تأكد من طريقة كتابتك لرقم السجل")
        if not help.is_number(self.customer_tax_num_e.get().strip()):
            errors.append("الرقم الجبائي يجب أن يتكون من أرقام فقط")
        if not help.is_number(self.customer_tax_item_e.get().strip()):
            errors.append("رقم مادة الضرائب يجب أن يتكون من أرقام فقط")
        if not help.is_number(self.customer_statistical_id_e.get().strip()):
            errors.append("رقم التعريف الاحصائي يجب أن يتكون من أرقام فقط".strip())
        return errors

    def clear_entries(self):
        self.selected_customer = None
        self.customer_name_e.delete(0, tk.END)
        self.customer_lastname_e.delete(0, tk.END)
        self.customer_phone_e.delete(0, tk.END)
        self.customer_address_e.delete(0, tk.END)
        self.customer_reg_id_e.delete(0, tk.END)
        self.customer_tax_num_e.delete(0, tk.END)
        self.customer_tax_item_e.delete(0, tk.END)
        self.customer_statistical_id_e.delete(0, tk.END)

    def select_customer(self):
        values = self.my_tree.item(self.my_tree.focus(), 'values')
        self.clear_entries()
        self.selected_customer = values[8]
        self.customer_name_e.insert(0, values[7])
        self.customer_lastname_e.insert(0, values[6])
        self.customer_phone_e.insert(0, values[5] if values[5] != "غير متوفر" else "")
        self.customer_address_e.insert(0, values[4])
        self.customer_reg_id_e.insert(0, values[3])
        self.customer_tax_num_e.insert(0, values[2])
        self.customer_tax_item_e.insert(0, values[1])
        self.customer_statistical_id_e.insert(0, values[0])

    def add_customer(self):
        errors = self.validation()
        if errors:
            messagebox.showerror("خطأ", "\n".join(errors), parent=self.parent)
        else:
            if any(customer[5] == self.customer_reg_id_e.get().strip() for customer in self.data):
                messagebox.showwarning("تحذير", "تأكد من رقم السجل لأنه موجود مسبقا في قاعدة البيانات",
                                       parent=self.parent)
            else:
                data = {
                    "name": self.customer_name_e.get().strip(),
                    "last_name": self.customer_lastname_e.get().strip(),
                    "phone": self.customer_phone_e.get().strip(),
                    "address": self.customer_address_e.get().strip(),
                    "commercial_register_number": self.customer_reg_id_e.get().strip(),
                    "tax_number": self.customer_tax_num_e.get().strip(),
                    "tax_item_number": self.customer_tax_item_e.get().strip(),
                    "statistical_id": self.customer_statistical_id_e.get().strip()
                }
                msg = db.add_customer(data)
                if msg.startswith("تم"):
                    self.update_customers_list()

                    messagebox.showinfo("", msg)
                else:
                    messagebox.showerror("", msg)

    def update_customers_list(self):
        self.my_tree.delete(*self.my_tree.get_children())
        self.data = db.get_customers()

        # Insert the updated data into the Treeview
        for customer in self.data:
            self.my_tree.insert(parent='', index='end', text='',
                                values=(customer[8], customer[7], customer[6], customer[5], customer[4],
                                        customer[3] if customer[3] != '' else "غير متوفر", customer[2],
                                        customer[1], customer[0]))

        self.clear_entries()

    def update_customer(self):
        errors = self.validation()
        if errors:
            messagebox.showerror("خطأ", "\n".join(errors), parent=self.parent)
        else:
            if any(customer[5] == self.customer_reg_id_e.get().strip()
                   and str(customer[0]) != self.selected_customer for customer in self.data):
                messagebox.showwarning("تحذير", "تأكد من رقم السجل لأنه موجود مسبقا في قاعدة البيانات",
                                       parent=self.parent)
            elif any(str(customer[0]) == self.selected_customer for customer in self.data):
                data = {
                    "name": self.customer_name_e.get().strip(),
                    "last_name": self.customer_lastname_e.get().strip(),
                    "phone": self.customer_phone_e.get().strip(),
                    "address": self.customer_address_e.get().strip(),
                    "commercial_register_number": self.customer_reg_id_e.get().strip(),
                    "tax_number": self.customer_tax_num_e.get().strip(),
                    "tax_item_number": self.customer_tax_item_e.get().strip(),
                    "statistical_id": self.customer_statistical_id_e.get().strip(),
                    "id": self.selected_customer.strip()
                }
                msg = db.update_customer(data)
                if msg.startswith("تم"):
                    self.update_customers_list()
                    messagebox.showinfo("", msg)
                else:
                    messagebox.showerror("", msg)
            else:
                messagebox.showerror("خطأ", "هذا الزبون لا يوجد في قاعدة البيانات", parent=self.parent)


class Bills:
    def __init__(self, parent, data):
        self.orders_ids = {}
        self.search = None
        self.checkbox = tk.IntVar()
        self.data = data
        self.my_tree = None
        self.parent = parent
        self.frame = ttk.Frame(parent)
        self.setup_ui()

    def setup_ui(self):
        tree_frame = ttk.Frame(self.frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        checkbox = ttk.Checkbutton(tree_frame, text="إظهار الفواتير التي لم يتم دفعها فقط", command=self.filter_results,
                                   variable=self.checkbox)
        checkbox.pack(padx=20, pady=10)

        self.search = ttk.Entry(tree_frame, justify="right", font=normal_text)
        self.search.pack(padx=20, pady=10, fill=tk.X)
        self.search.bind("<Return>", lambda event: self.update_search())

        self.my_tree = ttk.Treeview(tree_frame, selectmode="extended")
        self.my_tree.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        scroll_frame = ttk.Frame(tree_frame)
        scroll_frame.pack(side=tk.RIGHT, fill=tk.Y)

        tree_scroll = ttk.Scrollbar(scroll_frame, orient="vertical", command=self.my_tree.yview)
        tree_scroll.pack(side=tk.LEFT, fill="y")

        self.my_tree.config(yscrollcommand=tree_scroll.set)

        cols = ("state", "price", "date", "customer", "number", "id")
        self.my_tree['columns'] = cols

        self.my_tree.column("#0", width=0, stretch=tk.NO)
        self.my_tree.column("id", width=0, stretch=tk.NO)
        self.my_tree.column("number", anchor=tk.CENTER)
        self.my_tree.column("customer", anchor=tk.CENTER)
        self.my_tree.column("date", anchor=tk.CENTER)
        self.my_tree.column("price", anchor=tk.CENTER)
        self.my_tree.column("state", anchor=tk.CENTER)

        self.my_tree.heading("#0", text="", anchor=tk.W)
        self.my_tree.heading("number", anchor=tk.CENTER, text="الرقم")
        self.my_tree.heading("customer", anchor=tk.CENTER, text="الزبون")
        self.my_tree.heading("date", anchor=tk.CENTER, text="التاريخ")
        self.my_tree.heading("price", anchor=tk.CENTER, text="السعر الاجمالي")
        self.my_tree.heading("state", anchor=tk.CENTER, text="الحالة")
        self.my_tree.bind('<Double-1>',
                          lambda event: self.on_double_click()
                          if self.my_tree.identify_region(event.x, event.y) != "heading" else None)
        self.my_tree.bind('<Delete>',
                          lambda event: self.delete_row()
                          if self.my_tree.identify_region(event.x, event.y) != "heading" else None)
        self.update_bills()

    def delete_row(self):
        selected_items = self.my_tree.selection()
        if len(selected_items) > 1:
            messagebox.showerror("خطأ", "يمكنك حذف فاتورة واحدة فقط في كل مرة", parent=self.parent)
        else:
            msg = "هل أنت متأكد أنك تريد حذف هذه الفاتورة؟ لا يمكن التراجع عن هذا الإجراء."
            if messagebox.askyesno("تأكيد", msg, parent=self.parent):
                selected_item = selected_items[0]
                values = self.my_tree.item(selected_item, "values")
                selected_order = self.data[int(values[-1])]
                # print(selected_order)
                print(db.update_product_quantity(selected_order["products"], add=True))
                msg = db.delete_order(selected_order["id"])
                if msg.startswith("تم"):
                    messagebox.showinfo("نجاح", msg, parent=self.parent)
                else:
                    messagebox.showerror("خطأ", msg, parent=self.parent)
                self.update_bills()

    def filter_results(self):
        self.my_tree.delete(*self.my_tree.get_children())
        self.search.delete(0, tk.END)
        if self.checkbox.get():
            for o_id, record in self.data.items():
                if record["order_status"].startswith("لم"):
                    price = 0.0
                    for p in record["products"]:
                        price += ((float(p["price_at_order_time"]) * float(p["quantity"])) + float(
                            p["vat_tax_at_order_time"])
                                  + float(p["stamp_tax_at_order_time"]))
                    self.my_tree.insert(parent='', index='end', text='',
                                        values=(record["order_status"], price, record["order_date"],
                                                f"{record['customer_info']['name']} {record['customer_info']['last_name']}",
                                                self.orders_ids[o_id], o_id))
        else:
            for o_id, record in self.data.items():
                price = 0.0
                for p in record["products"]:
                    price += ((float(p["price_at_order_time"]) * float(p["quantity"])) + float(
                        p["vat_tax_at_order_time"])
                              + float(p["stamp_tax_at_order_time"]))
                self.my_tree.insert(parent='', index='end', text='',
                                    values=(record["order_status"], price, record["order_date"],
                                            f"{record['customer_info']['name']} {record['customer_info']['last_name']}",
                                            self.orders_ids[o_id], o_id))

    def on_double_click(self):
        item = self.my_tree.selection()[0]
        item_values = self.my_tree.item(item, "values")
        x, y, _, _ = self.my_tree.bbox(item)
        x += self.my_tree.winfo_rootx() + 50
        y += self.my_tree.winfo_rooty() + 20

        # Create a menu
        menu = tk.Menu(self.my_tree, tearoff=0, font=label_text, bg="maroon", fg="white", activebackground="#DD5746")

        menu.add_command(label="تحميل الفاتورة", command=lambda: self.download_bill(item_values[-1]))
        msg = "تم الدفع" if item_values[0].startswith("لم") else "لم يتم الدفع"
        menu.add_command(label=msg, command=lambda: (db.update_order_status(item_values[-1]), self.update_bills()))
        # Display the menu at the location of the double click event
        menu.post(x, y)

    def update_bills(self):
        self.data = db.get_orders()
        orders_ids_data = [(order_id, order_info["order_date"]) for order_id, order_info in self.data.items()]
        self.orders_ids = help.get_orders_ids(orders_ids_data)

        self.filter_results()

        # self.checkbox.set(0)
        #
        # self.my_tree.delete(*self.my_tree.get_children())
        # for o_id, record in self.data.items():
        #     price = 0.0
        #     for p in record["products"]:
        #         price += ((float(p["price_at_order_time"]) * float(p["quantity"])) + float(p["vat_tax_at_order_time"])
        #                   + float(p["stamp_tax_at_order_time"]))
        #     self.my_tree.insert(parent='', index='end', text='',
        #                         values=(record["order_status"], f"{price}", record["order_date"],
        #                                 f"{record['customer_info']['name']} {record['customer_info']['last_name']}",
        #                                 self.orders_ids[o_id], o_id))

    def download_bill(self, bill_id):
        # get info
        data = self.data[int(bill_id)]

        # Create a new workbook
        wb = load_workbook("template.xlsx")
        ws = wb.active

        # Define default font style
        default_font = Font(name='Arial', size=12)  # Set the desired font name and size
        label_font = Font(name='Arial', size=14, bold=True)
        title_font = Font(name='Arial', size=16, bold=True)
        # Apply default font style to all cells in the worksheet
        for row in ws.iter_rows():
            for cell in row:
                cell.font = default_font
                cell.alignment = Alignment(wrap_text=True)

        # date
        ws.merge_cells(f'A1:B1')
        order_date = datetime.strptime(data["order_date"], "%Y-%m-%d")
        ws['A1'] = f"عين التراب في : {order_date.strftime("%Y/%m/%d")}"

        # Supplier info
        ws['C2'] = "Jane Doe"
        ws['C3'] = "تجارة بالجملة"
        ws['C4'] = "........"
        ws['C5'] = "00/00 - 0000000 A 00"
        ws['C6'] = "0000000000"
        ws['C7'] = "0000000000"
        ws['C8'] = "0000000000"

        # customer info
        ws['A9'] = data["customer_info"]["name"] + " " + data["customer_info"]["last_name"]
        ws['A10'] = data["customer_info"]["address"]
        ws['A11'] = data["customer_info"]["commercial_register_number"]
        ws['A12'] = data["customer_info"]["tax_number"]
        ws['A13'] = data["customer_info"]["tax_item_number"]
        ws['A14'] = data["customer_info"]["statistical_id"]

        # bill number
        ws['A16'] = f"فاتورة رقم : {self.orders_ids[int(bill_id)]}"
        ws['A16'].font = title_font
        ws['A16'].alignment = Alignment(horizontal='center', vertical='center')

        # aligning
        cells = ["C2", "C4", "C5", "C6", "C7", "C8", "A9", "A10", "A11", "A12", "A13", "A14"]
        for cell in cells:
            ws[cell].alignment = Alignment(horizontal='right', vertical='center')

        cells = ["A18", "B18", "C18", "D18", "E18", "F18"]
        for cell in cells:
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            ws[cell].font = label_font

        i = 0
        tva_sum = 0
        stamp_price = 0
        price = 0
        while i < len(data["products"]):
            tva_sum += float(data["products"][i]["vat_tax_at_order_time"])
            stamp_price += float(data["products"][i]["stamp_tax_at_order_time"])
            price += float(data["products"][i]["quantity"]) * float(data["products"][i]["price_at_order_time"])

            # print to cells
            ws[f'F{19 + i}'] = i + 1
            ws[f'E{19 + i}'] = data["products"][i]["product_name"]
            ws[f'D{19 + i}'] = data["products"][i]["product_category"]
            ws[f'C{19 + i}'] = data["products"][i]["quantity"]
            ws[f'B{19 + i}'] = '{:.2f}'.format(data["products"][i]["price_at_order_time"])
            ws[f'A{19 + i}'] = '{:.2f}'.format(
                float(data["products"][i]["quantity"]) * float(data["products"][i]["price_at_order_time"]))

            for cell in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{cell}{19 + i}'].border = Border(left=Side(style="thin"))
                ws[f'{cell}{19 + i}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'F{19 + i}'].border = Border(right=Side(style="thin"), left=Side(style="thin"))
            ws[f'F{19 + i}'].alignment = Alignment(horizontal='center', vertical='center')
            i += 1

        for cell in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{cell}{19 + i}'].border = Border(top=Side(style="thin"))

        ws[f'A{19 + i}'] = '{:.2f}'.format(price)
        ws[f'A{19 + i + 1}'] = '{:.2f}'.format(tva_sum)
        ws[f'A{19 + i + 2}'] = '{:.2f}'.format(stamp_price)
        ws[f'A{19 + i + 3}'] = '{:.2f}'.format(price + tva_sum + stamp_price)

        ws[f'A{19 + i}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{19 + i + 1}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{19 + i + 2}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{19 + i + 3}'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(f'B{19 + i}:C{19 + i}')
        ws.merge_cells(f'B{19 + i + 1}:C{19 + i + 1}')
        ws.merge_cells(f'B{19 + i + 2}:C{19 + i + 2}')
        ws.merge_cells(f'B{19 + i + 3}:C{19 + i + 3}')

        ws[f'B{19 + i}'] = "المبلغ دون الرسم"
        ws[f'B{19 + i + 1}'] = "مبلغ الرسم على القيمة المضافة"
        ws[f'B{19 + i + 2}'] = "ضريبة الطابع"
        ws[f'B{19 + i + 3}'] = "المبلغ باحتساب كل الرسوم"

        for cell in ['A', 'B']:
            for j in range(4):
                ws[f'{cell}{19 + i + j}'].border = Border(top=Side(style="thin"),
                                                          left=Side(style="thin"),
                                                          right=Side(style="thin"),
                                                          bottom=Side(style="thin"))
        for j in range(4):
            ws[f'C{19 + i + j}'].border = Border(right=Side(style="thin"))

        ws.merge_cells(f'D{19 + i + 5}:F{19 + i + 5}')
        ws.merge_cells(f'D{19 + i + 6}:F{19 + i + 6}')

        ws[f'D{19 + i + 5}'] = "يوقف سند هذه الفاتورة على المبلغ الإجمالي"
        ws[f'D{19 + i + 6}'] = "كيفية الدفع"
        ws[f'D{19 + i + 5}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'D{19 + i + 6}'].alignment = Alignment(horizontal='right', vertical='center')

        ws[f'B{19 + i + 7}'] = "إمضاء و ختم"
        ws[f'B{19 + i + 7}'].font = Font(bold=True, underline='single')
        ws[f'B{19 + i + 7}'].alignment = Alignment(horizontal='center', vertical='center')

        desktop_dir = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

        # Ask the user for the file path
        file_name = f"Facture N {self.orders_ids[int(bill_id)].replace('/', '-')}"
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],
                                                 initialdir=desktop_dir, initialfile=file_name)

        # If the user cancels, return without saving
        if not file_path:
            return

        # Save the workbook to the chosen file path
        wb.save(file_path)
        messagebox.showinfo("حفظ الفاتورة", f"تم حفظ الفاتورة بنجاح في\n{file_path}")

    def update_search(self):
        search_term = self.search.get().strip()
        if len(search_term):
            filtered_results = {}
            for order_id, order_data in self.data.items():
                customer_name = order_data["customer_info"]["name"].lower()
                customer_last_name = order_data["customer_info"]["last_name"].lower()
                if search_term in customer_name or search_term in customer_last_name:
                    filtered_results[order_id] = order_data
            data = filtered_results

            self.checkbox.set(0)

            self.my_tree.delete(*self.my_tree.get_children())
            for o_id, record in data.items():
                price = 0.0
                for p in record["products"]:
                    price += ((float(p["price_at_order_time"]) * float(p["quantity"])) + float(
                        p["vat_tax_at_order_time"])
                              + float(p["stamp_tax_at_order_time"]))
                self.my_tree.insert(parent='', index='end', text='',
                                    values=(record["order_status"], f"{price}", record["order_date"],
                                            f"{record['customer_info']['name']} {record['customer_info']['last_name']}",
                                            self.orders_ids[o_id], o_id))
        else:
            self.update_bills()


class UserInfo:
    def __init__(self, parent, data):
        self.data = data
        self.parent = parent
        self.frame = ttk.Frame(parent)
        self.setup_ui()

    def setup_ui(self):
        self.frame.columnconfigure(0, weight=1)
        # Create labels to display user information
        ttk.Label(self.frame, text="معلومات الحساب", font=heading, justify="right").grid(row=0, column=0, sticky="e",
                                                                                         columnspan=2, pady=50, padx=30)

        row = 1
        for key, value in self.data.items():
            ttk.Label(self.frame, text=key, font=label_text).grid(row=row, column=1, sticky='e', padx=30)
            ttk.Label(self.frame, text=value, font=normal_text, anchor="e").grid(row=row, column=0, sticky='e', padx=30)
            row += 1

        ttk.Separator(self.frame, orient='horizontal').grid(row=row, column=0, sticky="ew", columnspan=2, pady=50,
                                                            padx=30)

        ttk.Label(self.frame, text="تحميل نسخة من البيانات", font=heading, justify="right").grid(row=row + 1, column=0,
                                                                                                 sticky="e",
                                                                                                 columnspan=2, padx=30,
                                                                                                 pady=0)
        msg = "يمكنك حفظ نسخة من قاعدة البيانات لاستعمالها في مكان آخر من هنا"
        ttk.Label(self.frame, text=msg, font=label_text).grid(row=row + 2, column=1, sticky='e', padx=30, pady=0)
        ttk.Button(self.frame, text="حفظ", style="danger.TButton", command=self.backup).grid(row=row + 2, column=0,
                                                                                             sticky='e', padx=30,
                                                                                             pady=0)

    def backup(self):
        try:
            desktop_dir = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

            # Ask the user for the file path
            file_path = filedialog.asksaveasfilename(defaultextension=".db", initialdir=desktop_dir,
                                                     filetypes=[("SQLite Database files", "*.db")],
                                                     initialfile="data")
            # If the user cancels, return without exporting
            if not file_path:
                return

            # Copy the database file to the specified location
            shutil.copy('data.db', file_path)

            # Show a success message
            tk.messagebox.showinfo("اكتمل التحميل", f"تم حفظ بيانات بنجاح في \n {file_path}",
                                   parent=self.parent)

        except Exception as e:
            tk.messagebox.showerror("خطأ", f"حدث خطأ أثناء عميلة حفظ البيانات", parent=self.parent)


if __name__ == "__main__":
    windll.shcore.SetProcessDpiAwareness(1)
    # Set locale to Arabic (Algerian)
    locale.setlocale(locale.LC_ALL, 'ar_DZ.UTF-8')

    heading = ("Segoe UI", 16, "bold")
    bold_text = ('Segoe UI', 12, 'bold')
    normal_text = ('Segoe UI', 12, 'normal')
    label_text = ('Segoe UI', 11, 'bold')

    # user info
    user_info = {
        "الاسم و اللقب": "Jane Doe",
        "العنوان": "Address",
        "رقم السجل التجاري": "00/00 - 0000000 A 00",
        "الرقم الجبائي": "00000000000",
        "رقم مادة الضرائب": "00000000000",
        "رقم التعريف الاحصائي": "00000000000"
    }

    # Database
    db = database.Database()
    db.create_tables()

    products = db.get_products()
    customers = db.get_customers()
    bills = db.get_orders()

    root = tk.Tk()
    root.geometry("1280x800")
    root.minsize(1200, 800)
    root.title("نظام الفواتير")
    root.iconbitmap("logo.ico")

    root.option_add('*Ttk*direction', 'rtl')  # تعيين الاتجاه لجميع عناصر ttk

    sv_ttk.set_theme("light")

    style = ttk.Style()

    style.configure('TButton', font=bold_text)
    style.configure('Treeview', font=normal_text, rowheight=40)
    style.configure('Treeview.Heading', font=bold_text, rowheight=50)
    style.configure('TLabel', font=label_text)
    style.configure('TCombobox', justify='right', font=normal_text, insertcolor="red")
    style.configure('TCheckbutton', font=label_text)
    style.configure('TNotebook', tabposition='ne')
    style.configure('TNotebook.Tab', font=heading, padding=[10, 10])
    style.configure('danger.TButton', font=bold_text, foreground="darkred")
    style.configure('success.TButton', font=bold_text, foreground="#135D66")
    style.configure('warning.TButton', font=bold_text, foreground="maroon")

    notebook = ttk.Notebook(root)
    notebook.pack(fill=tk.BOTH, expand=tk.TRUE)

    bills_frame = Bills(notebook, bills)
    orders_frame = Orders(notebook, products, customers, bills_frame)
    products_frame = Products(notebook, products, orders_frame)
    customers_frame = Customers(notebook, customers, orders_frame)

    # Add the instance of ProductManagementApp frame to the Notebook
    notebook.add(UserInfo(notebook, user_info).frame, text="بياناتي")
    notebook.add(bills_frame.frame, text="الفواتير")
    notebook.add(products_frame.frame, text="السلع")
    notebook.add(customers_frame.frame, text="الزبائن")
    notebook.add(orders_frame.frame, text="طلب جديد")

    notebook.select(4)


    def refresh(event):
        global notebook, bills_frame, orders_frame, products_frame, customers_frame
        if notebook.index("current") == 1:
            bills_frame.update_bills()
        elif notebook.index("current") == 2:
            products_frame.update_products_list()
        elif notebook.index("current") == 3:
            customers_frame.update_customers_list()
        elif notebook.index("current") == 4:
            orders_frame.update_products(db.get_products())
            orders_frame.update_customers(db.get_customers())


    notebook.bind("<<NotebookTabChanged>>", refresh)

    root.mainloop()
